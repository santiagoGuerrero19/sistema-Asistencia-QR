from flask import Flask, render_template, request, redirect, send_file, session
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import qrcode

app = Flask(__name__)

app.secret_key = 'clave_secreta_proyecto_asistencia'

ARCHIVO_EXCEL = 'asistencia.xlsx'
def formatear_excel():
    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    encabezados = ['Nombre', 'Código', 'Materia', 'Fecha', 'Hora']

    # Crea encabezados si no existen
    if hoja.max_row == 1 and hoja['A1'].value is None:
        hoja.append(encabezados)

    # Estilos
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="2563EB")
        celda.alignment = Alignment(horizontal="center")

    for columna in hoja.columns:
        max_largo = 0
        letra = columna[0].column_letter

        for celda in columna:
            if celda.value:
                max_largo = max(max_largo, len(str(celda.value)))

        hoja.column_dimensions[letra].width = max_largo + 4

    # Congelar encabezados
    hoja.freeze_panes = "A2"

    # Crear tabla si hay datos
    if hoja.max_row >= 2:
        rango_tabla = f"A1:E{hoja.max_row}"

        if "TablaAsistencia" not in hoja.tables:
            tabla = Table(displayName="TablaAsistencia", ref=rango_tabla)

            estilo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            tabla.tableStyleInfo = estilo
            hoja.add_table(tabla)
        else:
            hoja.tables["TablaAsistencia"].ref = rango_tabla

    libro.save(ARCHIVO_EXCEL)
    libro.close()

@app.route('/')
def inicio():
    return render_template('index.html')

@app.route('/registrar', methods=['POST'])
def registrar():
    nombre = request.form['nombre'].strip()
    codigo = request.form['codigo'].strip()
    materia = request.form['materia'].strip()

    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    hora_actual = datetime.now().strftime('%H:%M:%S')

    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    for fila in hoja.iter_rows(values_only=True):
        if fila[0] is None:
            continue

        codigo_guardado = str(fila[1]).strip()
        materia_guardada = str(fila[2]).strip().lower()
        fecha_guardada = str(fila[3]).strip()

        if (
            codigo_guardado == codigo and
            materia_guardada == materia.lower() and
            fecha_guardada == fecha_actual
        ):
            libro.close()
            return render_template('duplicado.html')

    hoja.append([nombre, codigo, materia, fecha_actual, hora_actual])

    libro.save(ARCHIVO_EXCEL)   
    libro.close()
    formatear_excel()

    return redirect('/exito')

@app.route('/exito')
def exito():
    return render_template('exito.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form['usuario']
        password = request.form['password']

        if usuario == 'KevinG' and password == '202120070010':
            session['admin'] = True
            return redirect('/admin')
        else:
            return render_template('login.html', error='Usuario o contraseña incorrectos')

    return render_template('login.html')

@app.route('/admin')
def admin():
    if not session.get('admin'):
        return redirect('/login')
    
    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    registros = []

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        registros.append({
            'nombre': fila[0],
            'codigo': fila[1],
            'materia': fila[2],
            'fecha': fila[3],
            'hora': fila[4]
        })

    libro.close()

    return render_template('admin.html', registros=registros)


@app.route('/descargar')
def descargar():
    if not session.get('admin'):
        return redirect('/login')

    return send_file(ARCHIVO_EXCEL, as_attachment=True)

@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect('/')

@app.route('/qr')
def mostrar_qr():

    url_formulario = 'https://sistema-asistencia-qr-3zp5.onrender.com/'

    qr = qrcode.make(url_formulario)
    qr.save('static/qr_asistencia.png')

    return render_template('qr.html', url_formulario=url_formulario)